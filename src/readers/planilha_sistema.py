from __future__ import annotations
import re
from pathlib import Path
from typing import Union
import openpyxl
import pandas as pd

_RE_PREFIXO = re.compile(r"^\d+\s*-\s*")
_RODAPE = {"total do dia", "t o t a l g e r a l", "total geral"}

def _limpar_nome(v):
    s = str(v).strip() if v else ""
    return _RE_PREFIXO.sub("", s).strip()

def _limpar_valor(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace("R$","").replace("\xa0","").replace(" ","")
    s = re.sub(r"[^\d,.\-]","",s)
    if "," in s and "." not in s: s = s.replace(",",".")
    elif "." in s and "," in s: s = s.replace(".","").replace(",",".")
    try: return float(s)
    except: return 0.0

def _eh_col_beneficiario(c):
    """Detecta coluna de beneficiário: Terceiro, Cliente, Favorecido etc."""
    return any(p in c for p in ("terceiro", "cliente", "favorecido", "benefici"))

def _eh_col_valor(c):
    """Detecta coluna de valor: Vlr. Nom, Vlr.Total, Valor Nom, Valor Total etc."""
    return any(p in c for p in ("vlr. nom", "vlr.nom", "valor nom", "vlr. tot", "vlr.tot", "valor tot", "vlr.total", "valor total"))

def _eh_col_data(c):
    return any(p in c for p in ("data pag", "vencimento", "emiss", "data liq")) or c == "data"

def ler_planilha_sistema(caminho: Union[str, Path]) -> pd.DataFrame:
    caminho = Path(caminho)
    if caminho.suffix.lower() in (".xlsm", ".xlsx", ".xls"): return _ler_excel(caminho)
    elif caminho.suffix.lower() == ".csv": return _ler_csv(caminho)
    raise ValueError(f"Formato não suportado: {caminho.suffix}")

def _ler_excel(caminho: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    # tenta aba ativa primeiro, depois todas as abas
    abas = [wb.active] + [wb[s] for s in wb.sheetnames if wb[s] != wb.active]
    for ws in abas:
        resultado = _ler_aba(ws)
        if resultado is not None:
            return resultado
    raise ValueError("Nenhum lançamento encontrado na planilha do sistema. Verifique se o arquivo contém colunas de Data, Terceiro/Cliente e Vlr. Nom/Valor.")

def _ler_aba(ws) -> pd.DataFrame | None:
    header_row = None
    col_data = col_benef = col_valor = col_doc = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rs = [str(v).strip().lower() if v else "" for v in row]
        tem_data  = any(_eh_col_data(c) for c in rs)
        tem_benef = any(_eh_col_beneficiario(c) for c in rs)
        tem_valor = any(_eh_col_valor(c) for c in rs)
        # precisa ter ao menos data+beneficiário OU data+valor para ser cabeçalho
        if tem_data and (tem_benef or tem_valor):
            header_row = i
            for j, c in enumerate(rs):
                if _eh_col_data(c) and col_data is None:   col_data  = j
                if _eh_col_beneficiario(c) and col_benef is None: col_benef = j
                if _eh_col_valor(c) and col_valor is None: col_valor = j
                if ("título" in c or "titulo" in c) and col_doc is None: col_doc = j
            break

    if header_row is None:
        return None

    registros = []
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not any(row): continue

        # data
        dv = row[col_data] if col_data is not None and col_data < len(row) else None
        if dv is None: continue

        # valor
        val = _limpar_valor(row[col_valor] if col_valor is not None and col_valor < len(row) else None)
        if val == 0.0: continue

        # beneficiário
        cr = row[col_benef] if col_benef is not None and col_benef < len(row) else None
        benef = _limpar_nome(cr) if cr is not None else ""
        if any(r in benef.lower() for r in _RODAPE): continue

        doc = str(row[col_doc]).strip() if col_doc is not None and col_doc < len(row) and row[col_doc] else ""

        registros.append({
            "data": dv,
            "descricao": benef,
            "debito": val,
            "credito": 0.0,
            "saldo": 0.0,
            "banco": "Sistema",
            "horario": None,
            "documento": doc,
        })

    if not registros:
        return None

    df = pd.DataFrame(registros)
    df["data"] = pd.to_datetime(df["data"], errors="coerce")
    df = df.dropna(subset=["data"])
    df["descricao"] = df["descricao"].astype(str).str.strip()
    return df[["data","banco","descricao","credito","debito","saldo","horario","documento"]].reset_index(drop=True)

def _ler_csv(caminho: Path) -> pd.DataFrame:
    for sep in (";",",","\t"):
        for enc in ("utf-8","latin-1","cp1252"):
            try:
                df = pd.read_csv(caminho, sep=sep, encoding=enc, dtype=str, on_bad_lines="skip")
                mapa = {}
                for col in df.columns:
                    cl = col.strip().lower()
                    if _eh_col_data(cl): mapa[col] = "data"
                    elif _eh_col_beneficiario(cl): mapa[col] = "descricao"
                    elif _eh_col_valor(cl): mapa[col] = "valor"
                    elif "titulo" in cl or "título" in cl: mapa[col] = "documento"
                df = df.rename(columns=mapa)
                if "data" in df.columns and ("descricao" in df.columns or "valor" in df.columns):
                    df["debito"] = df.get("valor", pd.Series([0.0]*len(df))).apply(_limpar_valor)
                    df["credito"] = 0.0; df["saldo"] = 0.0
                    df["banco"] = "Sistema"; df["horario"] = None
                    if "descricao" not in df.columns: df["descricao"] = ""
                    if "documento" not in df.columns: df["documento"] = ""
                    df["descricao"] = df["descricao"].apply(_limpar_nome)
                    df["data"] = pd.to_datetime(df["data"], dayfirst=True, errors="coerce")
                    df = df.dropna(subset=["data"])
                    df = df[df["debito"] > 0]
                    cols = ["data","banco","descricao","credito","debito","saldo","horario","documento"]
                    for c in cols:
                        if c not in df.columns: df[c] = ""
                    return df[cols].reset_index(drop=True)
            except: continue
    raise ValueError(f"Não foi possível ler CSV: {caminho}")
