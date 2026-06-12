"""
Controle de Débito — Sistema Interno v1
Tabela: Débitos do Sistema (planilha interna) × Débitos Efetivados (extratos bancários)
"""

import io, sys, tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent))

st.set_page_config(page_title="Controle de Débito — Sistema Interno",
                   page_icon="🌾", layout="wide")

BANCO_POR_NOME = {
    "bradesco_alimentos": "Bradesco Alimentos", "bradescalimentos": "Bradesco Alimentos",
    "bradescoalimentos": "Bradesco Alimentos", "bradesco alimentos": "Bradesco Alimentos",
    "bradesco": "Bradesco", "sicredi": "Sicredi",
    "banrisul": "Banrisul",
    "bb_alimentos": "BB Alimentos", "bbalimentos": "BB Alimentos",
    "bb": "BB",
}

_CONSULTA_BANRISUL = ("banrisul_consulta", "consulta_banrisul",
                      "banrisul consulta", "consulta banrisul",
                      "banrisul_operacoes", "operacoes_banrisul",
                      "banrisul operacoes", "operacoes banrisul",
                      "consulta_operacoes", "consulta operacoes")

def eh_consulta_banrisul(nome):
    n = nome.lower()
    if "banrisul" in n and "consulta" in n: return True
    return any(p in n for p in _CONSULTA_BANRISUL)

def detectar_banco(nome):
    n = nome.lower()
    for k, v in BANCO_POR_NOME.items():
        if k in n: return v
    return None

def ler_extrato(nome, conteudo):
    from src.readers.bradesco import LeitorBradesco
    from src.readers.sicredi  import LeitorSicredi
    from src.readers.bb       import LeitorBB
    from src.readers.banrisul import LeitorBanrisul
    LEITORES = {"Bradesco": LeitorBradesco, "Bradesco Alimentos": LeitorBradesco,
                "Sicredi": LeitorSicredi, "BB": LeitorBB, "BB Alimentos": LeitorBB,
                "Banrisul": LeitorBanrisul}
    banco = detectar_banco(nome)
    if not banco:
        return None, f"'{nome}' não reconhecido"
    suffix = Path(nome).suffix
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(conteudo); tmp_path = Path(tmp.name)
    try:
        df = LEITORES[banco](tmp_path).ler()
        df["banco"] = banco
        return df, None
    except Exception as e:
        return None, f"Erro ao ler '{nome}': {e}"
    finally:
        tmp_path.unlink(missing_ok=True)

def ler_planilha_sistema(conteudo, filename=".xlsm"):
    from src.readers.planilha_sistema import ler_planilha_sistema as _ler
    suffix = Path(filename).suffix or ".xlsm"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(conteudo); tmp_path = Path(tmp.name)
    try:
        return _ler(tmp_path), None
    except Exception as e:
        return None, f"Erro ao ler planilha do sistema: {e}"
    finally:
        tmp_path.unlink(missing_ok=True)


def conciliar(df_prev, df_banco, limite_alerta=1500.0):
    from difflib import SequenceMatcher
    import re
    STOP = {"nf","nota","fiscal","ltda","sa","eireli","me","epp","pag","pgto",
            "ted","pix","de","da","do","das","dos","ao","as","os",
            "boleto","pagamento","transferencia","transf",
            "e","a","o","em","na","no","por","com","um","uma",
            "log","transp","trans","logistica","comercio",
            "industria","servicos","alimentos","agro","agropecuaria",
            "transportes","distribuidora","distribuidores"}
    _norm_cache = {}
    def norm(s):
        if s not in _norm_cache:
            t = re.sub(r"[^\w\s]", " ", str(s).upper())
            _norm_cache[s] = " ".join(w for w in t.split() if w.lower() not in STOP)
        return _norm_cache[s]
    _KW_PREF = {"PREFEITURA","PREF","MUNICIPAL","MUNICIPIO"}
    _KW_IMP  = {"IPTU","ISS","ISSQN","ICMS","ALVARA","LICENCA","TRIBUTO",
                "CONTRIBUICAO","IMPOSTO","ITU","IPTR","TAXAS","TAXA"}
    def _eh_imp_pref(pa, pb): return bool(pb & _KW_PREF) and bool(pa & _KW_IMP)
    _ALIASES = [
        ({"VIVO"},{"TELEFONICA","TELEF"}),
        ({"CACISM","CAM"},{"CACISM","CAM","CAMARA","COMERCIO"}),
        ({"BORGES"},{"SAFE2PAY","SAFE"}),
    ]
    def _tem_alias(a, b):
        for ga, gb in _ALIASES:
            if (a&ga and b&gb) or (a&gb and b&ga): return True
        return False
    def _pfx(sa, sb, ml=4):
        c=0
        for a in sa:
            if len(a)<ml: continue
            for b in sb:
                if len(b)<ml: continue
                if a!=b and (b.startswith(a) or a.startswith(b)): c+=1
        return c
    def _comuns(a,b): return a&b, _pfx(a,b)
    def sim_nome(a,b):
        na,nb=norm(a),norm(b)
        if not na or not nb: return 0.0
        wa,wb_=set(na.split()),set(nb.split())
        ex,pf=_comuns(wa,wb_)
        tm=len(ex)+pf
        if tm: return 0.55+0.1*min(tm,3)
        return SequenceMatcher(None,na,nb).ratio()
    _OP=["TARIFA","TAXA","IOF","JUROS","INSS","FGTS","SALDO","RENTAB","FACILCRED","RENDE FACIL","DEBITO SERV"]
    _op_pat="|".join(_OP)
    deb_banco=df_banco[(df_banco["debito"]>0)&~df_banco["descricao"].str.upper().str.contains(_op_pat,regex=True,na=False)].copy().reset_index(drop=True)
    deb_prev=df_prev[df_prev["debito"]>0].copy().reset_index(drop=True)
    pn={ip:(norm(r["descricao"]),set(norm(r["descricao"]).split())) for ip,r in deb_prev.iterrows()}
    bn={ib:(norm(r["descricao"]),set(norm(r["descricao"]).split())) for ib,r in deb_banco.iterrows()}
    scores={}
    for ip,prev in deb_prev.iterrows():
        na,wa=pn[ip]
        for ib,deb in deb_banco.iterrows():
            dv=abs(deb["debito"]-prev["debito"])/max(prev["debito"],1)
            if dv>0.60: continue
            try: dd=abs((deb["data"]-prev["data"]).days)
            except: dd=999
            nb_,wb_=bn[ib]
            ex,pf=_comuns(wa,wb_)
            exs={w for w in ex if len(w)>=2}
            tm=len(exs)+pf
            if tm: sn=0.55+0.1*min(tm,3); ld=10
            elif _tem_alias(wa,wb_): sn=0.70; ld=10
            elif _eh_imp_pref(wa,wb_): sn=0.65; ld=10
            elif ex: sn=0.32; ld=5
            else: continue
            if dd>ld: continue
            sv=max(0.0,1-dv/0.60); sd=max(0.0,1-dd/6)
            scores[(ip,ib)]=sn*0.50+sv*0.35+sd*0.15
    usados_prev=set(); usados_banco=set(); atrib={}
    for (ip,ib),sc in sorted(scores.items(),key=lambda x:x[1],reverse=True):
        if ip in usados_prev or ib in usados_banco: continue
        if sc<0.30: break
        atrib[ip]=(ib,sc); usados_prev.add(ip); usados_banco.add(ib)
    livres=[(ib,deb_banco.loc[ib,"debito"],set(norm(deb_banco.loc[ib,"descricao"]).split()),norm(deb_banco.loc[ib,"descricao"])) for ib in deb_banco.index if ib not in usados_banco]
    atrib_multi={}
    for ip,prev in deb_prev.iterrows():
        if ip in atrib: continue
        pv=prev["debito"]; pnorm=norm(prev["descricao"]); pw=set(pnorm.split())
        def _cand(ws,bn_):
            if not ws and not bn_: return False
            if pw&ws: return True
            if _pfx(pw,ws)>0: return True
            if _tem_alias(pw,ws): return True
            if _eh_imp_pref(pw,ws): return True
            if pnorm and bn_ and SequenceMatcher(None,pnorm,bn_).ratio()>=0.65: return True
            return False
        cands=[(ib,v) for ib,v,ws,bn_ in livres if ib not in usados_banco and v<pv and _cand(ws,bn_)]
        if len(cands)<2: continue
        if sum(v for _,v in cands)<pv*0.70: continue
        cands.sort(key=lambda x:x[1],reverse=True)
        sel,tot=[],0.0
        for ib,v in cands:
            if tot+v<=pv*1.20: sel.append(ib); tot+=v
        if len(sel)>=2 and abs(tot-pv)/max(pv,1)<=0.20:
            atrib_multi[ip]=(sel,tot)
            for ib in sel: usados_banco.add(ib)
            usados_prev.add(ip)
    atrib_enr={}
    for ip,(ibp,sc) in list(atrib.items()):
        prev=deb_prev.loc[ip]; deb=deb_banco.loc[ibp]
        pv=prev["debito"]; pago=deb["debito"]
        if pago>=pv*0.99 or pago>pv: continue
        pnorm=norm(prev["descricao"]); pw=set(pnorm.split())
        extras=[(ib,v) for ib,v,ws,bn_ in livres if ib not in usados_banco and (ws or bn_) and v<=(pv-pago)*1.30 and (pw&ws or _pfx(pw,ws)>0 or _tem_alias(pw,ws) or _eh_imp_pref(pw,ws))]
        if not extras: continue
        extras.sort(key=lambda x:x[1],reverse=True)
        sel_e,acum=[],pago
        for ib,v in extras:
            if acum+v<=pv*1.20: sel_e.append(ib); acum+=v
        if not sel_e: continue
        tf=round(acum,2)
        if abs(tf-pv)/max(pv,1)<=0.20:
            atrib_enr[ip]=(ibp,sel_e,tf)
            for ib in sel_e: usados_banco.add(ib)
    linhas=[]
    _GEN={"PGTO","PAGO","BOLETO","DEBITO","CREDITO","TRANSFERENCIA","TED","PIX","PAGAMENTO"}
    def _mn(desc):
        ps=[p.strip() for p in str(desc).split("/")]
        for p in reversed(ps):
            if set(p.upper().split())-_GEN: return p
        return ps[0]
    def _lm(prev,debs,aibs,tot):
        diff=round(tot-prev["debito"],2); pct=diff/prev["debito"]*100 if prev["debito"] else 0
        al=f"🔴 Diferença de R$ {abs(diff):,.2f}" if abs(diff)>0.01 else ""
        st_=("✅ OK (múlt.)" if abs(diff)<=prev["debito"]*0.02 else "⚠️ VALOR DIFERENTE (múlt.)")
        bs=debs["banco"].iloc[0] if debs["banco"].nunique()==1 else "Múltiplos"
        bn__=""
        for d in debs["descricao"]:
            c=_mn(d)
            if c and set(c.upper().split())-_GEN: bn__=c; break
        if not bn__: bn__=debs["descricao"].iloc[0]
        return {"Status":st_,"🔴 Alerta":al,"Data Prevista":prev["data"],"Beneficiário Previsto":prev["descricao"],
                "Valor Previsto (R$)":prev["debito"],"Data Pago":debs["data"].min(),"Banco":bs,
                "Pago Para":f"Múltiplos boletos ({len(aibs)}x) / {bn__}",
                "Valor Pago (R$)":tot,"Diferença (R$)":diff,"Diferença (%)":round(pct,1)}
    for ip,prev in deb_prev.iterrows():
        if ip in atrib_enr:
            ibp,ie,tot=atrib_enr[ip]; aibs=[ibp]+ie; debs=deb_banco.loc[aibs]
            linhas.append(_lm(prev,debs,aibs,tot)); continue
        if ip in atrib_multi:
            ibs,tot=atrib_multi[ip]; debs=deb_banco.loc[ibs]
            linhas.append(_lm(prev,debs,ibs,round(tot,2))); continue
        if ip in atrib:
            ib,sc=atrib[ip]; deb=deb_banco.loc[ib]
            diff=round(deb["debito"]-prev["debito"],2); pct=diff/prev["debito"]*100 if prev["debito"] else 0
            sn=sim_nome(prev["descricao"],deb["descricao"])
            if abs(diff)<=prev["debito"]*0.02 and sn>=0.40: st_="✅ OK"
            elif abs(diff)>prev["debito"]*0.02 and sn>=0.40: st_="⚠️ VALOR DIFERENTE"
            elif abs(diff)<=prev["debito"]*0.02 and sn<0.40: st_="⚠️ BENEFICIÁRIO DIFERENTE"
            else: st_="⚠️ DIVERGÊNCIA"
            al=f"🔴 Diferença de R$ {abs(diff):,.2f}" if abs(diff)>0.01 else ""
            linhas.append({"Status":st_,"🔴 Alerta":al,"Data Prevista":prev["data"],
                           "Beneficiário Previsto":prev["descricao"],"Valor Previsto (R$)":prev["debito"],
                           "Data Pago":deb["data"],"Banco":deb["banco"],"Pago Para":deb["descricao"],
                           "Valor Pago (R$)":deb["debito"],"Diferença (R$)":diff,"Diferença (%)":round(pct,1)})
        else:
            linhas.append({"Status":"🕐 NÃO PAGO","🔴 Alerta":"","Data Prevista":prev["data"],
                           "Beneficiário Previsto":prev["descricao"],"Valor Previsto (R$)":prev["debito"],
                           "Data Pago":None,"Banco":"","Pago Para":"","Valor Pago (R$)":None,
                           "Diferença (R$)":-prev["debito"],"Diferença (%)":-100.0})
    OPER=["TARIFA","TAXA","IOF","JUROS","INSS","FGTS","SALDO","RENTAB","FACILCRED","RENDE FACIL","DEBITO SERV"]
    for ib,deb in deb_banco.iterrows():
        if ib in usados_banco: continue
        if any(p in str(deb["descricao"]).upper() for p in OPER): continue
        linhas.append({"Status":"🚨 NÃO PREVISTO","🔴 Alerta":"","Data Prevista":None,
                       "Beneficiário Previsto":"","Valor Previsto (R$)":None,
                       "Data Pago":deb["data"],"Banco":deb["banco"],"Pago Para":deb["descricao"],
                       "Valor Pago (R$)":deb["debito"],"Diferença (R$)":deb["debito"],"Diferença (%)":None})
    return pd.DataFrame(linhas)


def cor_linha(row):
    s=str(row.get("Status",""))
    if s.startswith("✅"): bg="#d4edda; color:#155724"
    elif "BENEFICIÁRIO" in s: bg="#f8d7da; color:#721c24"
    elif s.startswith("⚠️"): bg="#fff3cd; color:#856404"
    elif s.startswith("🚨"): bg="#f8d7da; color:#721c24"
    else: bg="#e2e3e5; color:#383d41"
    return [f"background-color:{bg}"]*len(row)

def fmt_brl(v):
    try: return f"R$ {float(v):,.2f}"
    except: return ""
def fmt_dt(v):
    try: return pd.Timestamp(v).strftime("%d/%m/%Y")
    except: return ""

st.title("🌾 Débitos Sistema × Efetivados")
st.caption("Moinho de Trigo — comparação: o que o sistema registrou vs o que saiu dos bancos")
with st.sidebar:
    st.header("📂 Arquivos")
    extratos_up=st.file_uploader("Extratos bancários (PDF, XLSX, CSV)",type=["pdf","xlsx","xls","csv"],accept_multiple_files=True)
    planilha_up=st.file_uploader("Planilha do sistema (XLSM, XLSX, CSV)",type=["xlsm","xlsx","xls","csv"])
    st.caption("Extratos: bradesco_*.pdf · sicredi_*.pdf · bb_*.pdf · banrisul_*.pdf")
    st.divider()
    data_sel=st.date_input("📅 Data",value=datetime.now().date())
    mostrar_periodo=st.toggle("Ver período",value=False)
    if mostrar_periodo: data_fim=st.date_input("Até",value=datetime.now().date())
    else: data_fim=data_sel
    rodar=st.button("▶ Comparar",type="primary",use_container_width=True)
if not rodar:
    st.markdown("""### Como funciona\n1. Upload dos **extratos bancários** (PDF/Excel)\n2. Upload da **planilha do sistema** (.xlsm/.xlsx)\n3. Selecione a **data** e clique **▶ Comparar**\n\n| Cor | Significado |\n|---|---|\n| 🟢 Verde | Pago conforme sistema |\n| 🟡 Amarelo | Valor diferente |\n| 🔴 Vermelho | Saiu do banco sem estar no sistema |\n| ⬜ Cinza | No sistema, ainda não pago |""")
    st.stop()
if not extratos_up: st.error("Faça upload dos extratos."); st.stop()
if not planilha_up: st.error("Faça upload da planilha do sistema."); st.stop()
with st.spinner("Lendo extratos..."):
    dfs=[]; dfs_c=[]
    for f in extratos_up:
        c=f.read()
        if eh_consulta_banrisul(f.name):
            from src.readers.banrisul_consulta import ler_consulta_banrisul
            with tempfile.NamedTemporaryFile(suffix=Path(f.name).suffix,delete=False) as tmp:
                tmp.write(c); tp=Path(tmp.name)
            try:
                dc=ler_consulta_banrisul(tp)
                if not dc.empty: dfs_c.append(dc); st.sidebar.success(f"✔ {f.name} (consulta Banrisul)")
            except Exception as e: st.warning(str(e))
            finally: tp.unlink(missing_ok=True)
        else:
            df,err=ler_extrato(f.name,c)
            if err: st.warning(err)
            else: dfs.append(df); st.sidebar.success(f"✔ {f.name} → {df['banco'].iloc[0]}")
if not dfs: st.error("Nenhum extrato lido."); st.stop()
df_banco=pd.concat(dfs,ignore_index=True)
if dfs_c:
    df_cons=pd.concat(dfs_c,ignore_index=True)
    lv={}
    for _,row in df_cons.iterrows(): lv.setdefault(round(row["valor"],2),[]).append((row["data"],row["beneficiario"]))
    _GBN={"pgto boleto","pag boleto","pagamento boleto","debito automatico","arrecadacao","cobranca","debito transferencia","transferencia","ted","pix"}
    def _enr(row):
        if row["banco"]!="Banrisul": return row["descricao"]
        if not any(g in str(row["descricao"]).lower() for g in _GBN): return row["descricao"]
        v=round(float(row["debito"]),2)
        if v==0: return row["descricao"]
        cands=lv.get(v,[])
        if not cands: return row["descricao"]
        for d in (0,1,-1):
            alvo=row["data"]+pd.Timedelta(days=d)
            for i,(dc,nm) in enumerate(cands):
                if dc==alvo: cands.pop(i); return f"{row['descricao']} / {nm}"
        return row["descricao"]
    df_banco["descricao"]=df_banco.apply(_enr,axis=1)
df_banco_full=df_banco.copy()
with st.spinner("Lendo planilha do sistema..."):
    df_prev,err=ler_planilha_sistema(planilha_up.read(),filename=planilha_up.name)
    if err: st.error(err); st.stop()
    st.sidebar.success(f"✔ {planilha_up.name} ({len(df_prev)} lançamentos)")
with st.sidebar.expander("🔍 Amostra da planilha",expanded=False):
    st.dataframe(df_prev.head(5))
di=pd.Timestamp(data_sel); df_=pd.Timestamp(data_fim)
df_banco=df_banco[(df_banco["data"]>=di)&(df_banco["data"]<=df_)].copy()
df_prev=df_prev[(df_prev["data"]>=di)&(df_prev["data"]<=df_)].copy()
if df_banco.empty and df_prev.empty:
    st.warning(f"Nenhum lançamento para {data_sel.strftime('%d/%m/%Y')}."); st.stop()
pl=(data_sel.strftime("%d/%m/%Y") if data_sel==data_fim else f"{data_sel.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}")
st.info(f"📅 {pl} — {len(df_banco[df_banco['debito']>0])} débitos bancários · {len(df_prev[df_prev['debito']>0])} no sistema")
st.subheader("💰 Saldos dos Bancos")
bp=[b for b in ["Bradesco","Bradesco Alimentos","Sicredi","BB","BB Alimentos","Banrisul"] if not df_banco_full[df_banco_full["banco"]==b].empty]
cols=st.columns(max(len(bp),1))
for i,b in enumerate(bp):
    sub=df_banco_full[df_banco_full["banco"]==b]; snz=sub["saldo"][sub["saldo"]!=0]
    cols[i].metric(b,fmt_brl(float(snz.iloc[-1]) if not snz.empty else 0.0))
with st.spinner("Comparando..."):
    df_res=conciliar(df_prev,df_banco)
st.subheader("📊 Resumo")
nok=df_res["Status"].str.startswith("✅").sum(); ndiv=df_res["Status"].str.startswith("⚠️").sum()
nnao=df_res["Status"].str.startswith("🚨").sum(); npend=df_res["Status"].str.startswith("🕐").sum()
vdiv=df_res[df_res["Status"].str.startswith("⚠️")]["Diferença (R$)"].abs().sum()
vnao=df_res[df_res["Status"].str.startswith("🚨")]["Valor Pago (R$)"].sum()
c1,c2,c3,c4=st.columns(4)
c1.metric("✅ Conforme sistema",nok); c2.metric("⚠️ Valor diferente",ndiv,delta=fmt_brl(vdiv) if ndiv else None,delta_color="inverse")
c3.metric("🚨 Fora do sistema",nnao,delta=fmt_brl(vnao) if nnao else None,delta_color="inverse"); c4.metric("🕐 Não pago",npend)
st.subheader("📋 Tabela Comparativa")
cf1,cf2,cf3=st.columns(3)
so=sorted(df_res["Status"].unique()); ss=cf1.multiselect("Status",so,default=so)
bo=["Todos"]+sorted(df_banco["banco"].unique()); bs=cf2.selectbox("Banco",bo)
busca=cf3.text_input("Buscar nome")
dv=df_res[df_res["Status"].isin(ss)].copy()
if bs!="Todos": dv=dv[dv["Banco"]==bs]
if busca: dv=dv[dv["Beneficiário Previsto"].str.contains(busca,case=False,na=False)|dv["Pago Para"].str.contains(busca,case=False,na=False)]
ds=dv.copy()
ds["Data Prevista"]=ds["Data Prevista"].apply(fmt_dt); ds["Data Pago"]=ds["Data Pago"].apply(fmt_dt)
ds["Valor Previsto (R$)"]=ds["Valor Previsto (R$)"].apply(fmt_brl); ds["Valor Pago (R$)"]=ds["Valor Pago (R$)"].apply(fmt_brl)
ds["Diferença (R$)"]=ds["Diferença (R$)"].apply(lambda v:f"+{fmt_brl(v)}" if isinstance(v,(int,float)) and v>0 else (fmt_brl(v) if isinstance(v,(int,float)) else ""))
ds["Diferença (%)"]=ds["Diferença (%)"].apply(lambda v:f"{v:+.1f}%" if isinstance(v,(int,float)) else "")
st.dataframe(ds.style.apply(cor_linha,axis=1),use_container_width=True,height=620,
    column_config={"Status":st.column_config.TextColumn(width=200),"🔴 Alerta":st.column_config.TextColumn(width=280),
        "Data Prevista":st.column_config.TextColumn("Data Sist.",width=100),
        "Beneficiário Previsto":st.column_config.TextColumn("Previsto (Sistema)",width=230),
        "Valor Previsto (R$)":st.column_config.TextColumn("Vlr Sistema",width=130),
        "Data Pago":st.column_config.TextColumn("Data Pago",width=100),"Banco":st.column_config.TextColumn(width=90),
        "Pago Para":st.column_config.TextColumn(width=230),"Valor Pago (R$)":st.column_config.TextColumn("Vlr Pago",width=130),
        "Diferença (R$)":st.column_config.TextColumn("Diferença R$",width=120),
        "Diferença (%)":st.column_config.TextColumn("Diferença %",width=100)})
buf=io.BytesIO()
with pd.ExcelWriter(buf,engine="openpyxl") as w:
    ds.to_excel(w,sheet_name="Comparativo",index=False)
    df_res[df_res["Status"].str.startswith("🚨")].to_excel(w,sheet_name="Fora do Sistema",index=False)
    df_res[df_res["Status"].str.startswith("⚠️")].to_excel(w,sheet_name="Divergências",index=False)
    df_banco.to_excel(w,sheet_name="Extrato Consolidado",index=False)
import openpyxl as _ox; buf.seek(0); _wb=_ox.load_workbook(buf)
from openpyxl.styles import PatternFill,Font,Alignment
for ws_ in _wb.worksheets:
    for cell in ws_[1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2C3E50")
    ws_.freeze_panes="A2"
buf2=io.BytesIO(); _wb.save(buf2)
st.download_button("📥 Baixar Excel",data=buf2.getvalue(),file_name=f"comparativo_sistema_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
